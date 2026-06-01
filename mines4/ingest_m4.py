"""
ingest_m4.py — Ingest all three workbooks into m4_* tables (schema_v4.sql)
===========================================================================

Handles:
  • Mining Investment Model Prototype (4.20.2026).xlsx
      Source: Mine_Profile sheet (column C inputs only)
      Result: 1 mine, 1 commodity (Gold), 1 scenario (Single)

  • Mine1.xlsx  (Chinaka Resource Mining 3, License 12891L)
      Source: REE, Monazite, Spodumene sheets  +  REE Scenario Analysis
      Result: 1 mine, 3 commodities, Bear/Base/Bull for each commodity

  • mine11.xlsx  (M'Gomo Mine, License 9015L)
      Source: Gold sheet, Graphite Bear/Base/Bull sheets
      Result: 1 mine, 2 commodities (Gold=Single, Graphite=Bear/Base/Bull)

INPUT ROWS INGESTED (R = row number):
  R5  — production (t/yr per column year)
  R8  — commodity price (escalating per year)
  R14 — operating costs ($ p.a.)
  R19 — depreciation ($ p.a.)
  R24 — capex ($)

CALCULATED rows (R9-R11, R15-R16, R20-R21, R25-R28) are NOT ingested.

USAGE:
  pip install openpyxl supabase python-dotenv
  python ingest_m4.py --dry-run
  python ingest_m4.py
  python ingest_m4.py --json out.json
"""
from __future__ import annotations
import argparse, json, os, re, sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl not installed — run: pip install openpyxl")

BASE_DIR = Path(__file__).parent.parent / "mines"

SUPABASE_URL = os.environ.get("SUPABASE_URL",
    "https://snbnqwrxvptrfjsecljd.supabase.co")
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_KEY",
    os.environ.get("SUPABASE_ANON_KEY",
    "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9."
    "eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InNuYm5xd3J4dnB0cmZqc2VjbGpkIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3OTg5MzQ4MSwiZXhwIjoyMDk1NDY5NDgxfQ."
    "t92K9HW0uQpCWy08c6CPtGKynHN5ET3ymcfcupNJTO0"))


def _sb():
    from supabase import create_client
    return create_client(SUPABASE_URL, SUPABASE_KEY)


# ── Utility ───────────────────────────────────────────────────────────────────

def _n(v) -> Optional[float]:
    try: return float(v) if v is not None else None
    except: return None

def _i(v) -> Optional[int]:
    f = _n(v)
    return int(round(f)) if f is not None else None

def _s(v) -> Optional[str]:
    return str(v).strip() if v is not None else None


def _cell(ws, row: int, col: int):
    v = ws.cell(row=row, column=col).value
    return None if (v is None or str(v).startswith("=")) else v


def _year_cols(ws) -> Dict[int, int]:
    """Return {year_number: col_index} from row 3 headers (Yr 0, Yr 1 ...)."""
    cols = {}
    for col in range(2, min((ws.max_column or 30) + 1, 50)):
        hdr = _s(_cell(ws, 3, col)) or ""
        m = re.search(r"(\d+)", hdr)
        if m and ("yr" in hdr.lower() or "year" in hdr.lower()):
            cols[int(m.group(1))] = col
    return cols


def _read_input_row(ws, row_num: int, year_cols: Dict[int, int]) -> Dict[int, Optional[float]]:
    """Return {year: value} for one input row across all year columns."""
    return {yr: _n(_cell(ws, row_num, col)) for yr, col in year_cols.items()}


def _derive_scalar(year_vals: Dict[int, Optional[float]],
                   is_price: bool = False,
                   is_capex: bool = False,
                   is_opex: bool = False,
                   is_prod: bool = False) -> dict:
    """Derive scalar inputs from year-by-year data."""
    result: dict = {}
    vals = {yr: v for yr, v in year_vals.items() if v is not None}

    if is_prod:
        nonzero = {yr: v for yr, v in vals.items() if v > 0}
        if nonzero:
            result["annual_production"] = max(nonzero.values())
            result["production_start_year"] = min(nonzero.keys())
        return result

    if is_price:
        sorted_yrs = sorted(vals.keys())
        nonzero = [vals[y] for y in sorted_yrs if vals[y] and vals[y] > 0]
        if nonzero:
            result["price_base"] = nonzero[0]
            if len(nonzero) >= 2:
                try:
                    n = len(nonzero) - 1
                    esc = (nonzero[-1] / nonzero[0]) ** (1 / n) - 1
                    result["price_escalation_rate"] = round(esc, 4) if abs(esc) < 0.20 else 0.0
                except: result["price_escalation_rate"] = 0.0
        return result

    if is_capex:
        # Initial capex: first large negative value
        sorted_yrs = sorted(vals.keys())
        for yr in sorted_yrs[:3]:
            v = vals.get(yr)
            if v is not None and v < 0:
                result["initial_capex"] = abs(v)
                result["capex_deployment_year"] = yr
                break
        # Sustaining: median of smaller negative values after initial
        sust = [abs(v) for yr, v in sorted(vals.items())
                if v is not None and v < 0 and yr > result.get("capex_deployment_year", -1)
                and abs(v) < result.get("initial_capex", 1e15) * 0.5]
        result["sustaining_capex_pa"] = sorted(sust)[len(sust)//2] if sust else 0.0
        return result

    if is_opex:
        nonzero = sorted(abs(v) for v in vals.values() if v and abs(v) > 0)
        if nonzero:
            result["opex_steady_state"] = nonzero[len(nonzero)//2]
            if len(nonzero) >= 2:
                try:
                    n = len(nonzero) - 1
                    esc = (nonzero[-1] / nonzero[0]) ** (1 / n) - 1
                    result["opex_escalation_rate"] = round(esc, 4) if 0 <= esc < 0.15 else 0.0
                except: result["opex_escalation_rate"] = 0.0
        return result

    return result


def _derive_dep(year_vals: Dict[int, Optional[float]]) -> Optional[float]:
    nonzero = sorted(abs(v) for v in year_vals.values() if v and abs(v) > 0)
    return nonzero[len(nonzero)//2] if nonzero else None


# ═══════════════════════════════════════════════════════════════════════════════
# PROTOTYPE  (Mine_Profile sheet → single commodity)
# ═══════════════════════════════════════════════════════════════════════════════

# Mine_Profile column C row map — VERIFIED against actual file
_PROFILE_ROWS = [
    (4,  3, "mine_name"),
    (6,  3, "country"),
    (7,  3, "license_number"),
    (8,  3, "mine_type"),
    (13, 3, "ore_reserve"),
    (13, 2, "reserve_unit"),
    (14, 3, "throughput_pa"),
    (15, 3, "life_of_mine_yr"),
    (18, 3, "commodity"),
    (19, 3, "price_base"),
    (20, 3, "grade"),
    (20, 2, "grade_unit"),
    (25, 3, "initial_capex"),
    (28, 3, "opex_steady_state"),
    (31, 3, "opex_escalation_rate"),
    (32, 3, "avg_depreciation_years"),
    (35, 3, "ramp_up_y1"),
    (36, 3, "ramp_up_y2"),
    (37, 3, "ramp_up_y3"),
    (40, 3, "tax_rate"),
    (41, 3, "royalty_rate"),
    (42, 3, "debt_funding"),
    (43, 3, "debt_term"),
    (44, 3, "interest_rate"),
    (45, 3, "wacc"),
    (46, 3, "closure_rehab_cost"),
]

_NUMERIC_FIELDS = {
    "ore_reserve","throughput_pa","life_of_mine_yr","price_base","grade",
    "initial_capex","opex_steady_state","opex_escalation_rate","avg_depreciation_years",
    "ramp_up_y1","ramp_up_y2","ramp_up_y3","tax_rate","royalty_rate",
    "debt_funding","debt_term","interest_rate","wacc","closure_rehab_cost",
}

def parse_prototype(wb_path: str) -> dict:
    wb = openpyxl.load_workbook(wb_path, data_only=True)
    ws = wb["Mine_Profile"]
    row = {}
    for r, c, field in _PROFILE_ROWS:
        v = ws.cell(row=r, column=c).value
        if v is None: continue
        row[field] = (_n(v) if field in _NUMERIC_FIELDS else _s(v))

    mine = {
        "mine_name":          row.get("mine_name", "Unknown"),
        "license_number":     row.get("license_number"),
        "country":            row.get("country", "Mozambique"),
        "mine_type":          row.get("mine_type"),
        "ore_reserve":        row.get("ore_reserve"),
        "reserve_unit":       row.get("reserve_unit", "m3"),
        "throughput_pa":      row.get("throughput_pa"),
        "life_of_mine_yr":    _i(row.get("life_of_mine_yr")),
        "wacc":               row.get("wacc"),
        "tax_rate":           row.get("tax_rate"),
        "royalty_rate":       row.get("royalty_rate"),
        "ramp_up_y1":         row.get("ramp_up_y1"),
        "ramp_up_y2":         row.get("ramp_up_y2"),
        "ramp_up_y3":         row.get("ramp_up_y3"),
        "closure_rehab_cost": row.get("closure_rehab_cost"),
        "debt_funding":       row.get("debt_funding"),
        "debt_term":          _i(row.get("debt_term")),
        "interest_rate":      row.get("interest_rate"),
        "source_file":        os.path.basename(wb_path),
    }

    commodity = row.get("commodity", "Unknown")
    scenario = {
        "scenario":              "Single",
        "price_base":            row.get("price_base"),
        "price_unit":            "$/kg" if commodity == "Gold" else "$/t",
        "initial_capex":         row.get("initial_capex"),
        "opex_steady_state":     row.get("opex_steady_state"),
        "price_escalation_rate": 0.0,
        "opex_escalation_rate":  row.get("opex_escalation_rate", 0.0),
        "avg_depreciation_years": _i(row.get("avg_depreciation_years")),
        "production_start_year": 1,
        "capex_deployment_year": 0,
        "basis_notes":           f"Prototype model — {commodity}",
    }

    # Derive annual_production from throughput × grade; also store grade for run_dcf
    tp = row.get("throughput_pa") or 0
    grade = row.get("grade") or 0
    if tp and grade:
        scenario["annual_production"]    = tp * grade / 1000.0  # m3 × g/m3 → kg
        scenario["avg_recovered_grade"]  = grade

    return {
        "source_file": os.path.basename(wb_path),
        "mine": mine,
        "commodities": [{
            "commodity":   commodity,
            "is_primary":  True,
            "has_scenarios": False,
            "scenarios":   [scenario],
        }],
    }


# ═══════════════════════════════════════════════════════════════════════════════
# MINE1  (Chinaka — REE/Monazite/Spodumene + Scenario Analysis)
# ═══════════════════════════════════════════════════════════════════════════════

# Scenario Analysis prices (from REE Scenario Analysis sheet)
_MINE1_PRICES = {
    "Bear": {"REE": 5000,  "Monazite": 800,  "Spodumene": 750,  "price_esc": 0.01},
    "Base": {"REE": 9000,  "Monazite": 1500, "Spodumene": 950,  "price_esc": 0.02},
    "Bull": {"REE": 13000, "Monazite": 2500, "Spodumene": 1200, "price_esc": 0.025},
}
# OPEX per commodity (from Exec Summary row 14, same for all scenarios)
_MINE1_OPEX = {"REE": 212e6, "Monazite": 31.8e6, "Spodumene": 21.2e6}
# CAPEX per commodity
_MINE1_CAPEX = {"REE": 1370e6, "Monazite": 79e6, "Spodumene": 50e6}
_MINE1_SUST  = {"REE": 55e6,   "Monazite": 3.2e6, "Spodumene": 2.0e6}
_MINE1_CLOSE = {"REE": 110e6,  "Monazite": 6.3e6, "Spodumene": 4.0e6}
_MINE1_PRICE_UNITS = {"REE": "$/t conc", "Monazite": "$/t", "Spodumene": "$/t"}


def _parse_mine1_dcf_sheet(ws) -> dict:
    """Parse one commodity sheet — extract ONLY input rows."""
    yc = _year_cols(ws)
    prod_d   = _read_input_row(ws, 5,  yc)
    price_d  = _read_input_row(ws, 8,  yc)
    opex_d   = _read_input_row(ws, 14, yc)
    dep_d    = _read_input_row(ws, 19, yc)
    capex_d  = _read_input_row(ws, 24, yc)

    prod_sc   = _derive_scalar(prod_d,  is_prod=True)
    price_sc  = _derive_scalar(price_d, is_price=True)
    capex_sc  = _derive_scalar(capex_d, is_capex=True)
    opex_sc   = _derive_scalar(opex_d,  is_opex=True)
    dep_pa    = _derive_dep(dep_d)

    years_input = []
    for yr in sorted(yc.keys()):
        years_input.append({
            "year":            yr,
            "production":      prod_d.get(yr),
            "commodity_price": price_d.get(yr),
            "operating_costs": opex_d.get(yr),
            "depreciation":    dep_d.get(yr),
            "capex":           capex_d.get(yr),
        })

    return {**prod_sc, **price_sc, **capex_sc, **opex_sc,
            "depreciation_pa": dep_pa,
            "years_input": years_input}


def parse_mine1(wb_path: str) -> dict:
    wb = openpyxl.load_workbook(wb_path, data_only=True)

    commodities = []
    comm_order = ["REE", "Monazite", "Spodumene"]

    for ci, comm in enumerate(comm_order):
        ws = wb[comm]
        base_inputs = _parse_mine1_dcf_sheet(ws)

        scenarios = []
        for scen_name, prices in _MINE1_PRICES.items():
            scen = {
                "scenario":              scen_name,
                "sheet_name":            f"{comm} ({scen_name})",
                "price_base":            prices[comm],
                "price_unit":            _MINE1_PRICE_UNITS[comm],
                "price_escalation_rate": prices["price_esc"],
                "annual_production":     base_inputs.get("annual_production"),
                "production_unit":       "tonnes",
                "opex_steady_state":     _MINE1_OPEX.get(comm),
                "opex_escalation_rate":  0.025,
                "initial_capex":         _MINE1_CAPEX.get(comm),
                "sustaining_capex_pa":   _MINE1_SUST.get(comm),
                "capex_deployment_year": 0,
                "depreciation_pa":       base_inputs.get("depreciation_pa"),
                "production_start_year": base_inputs.get("production_start_year", 1),
                "royalty_rate":          0.05,
                "basis_notes":           f"{comm} {scen_name} — REE ${prices['REE']}/t",
                "years_input":           base_inputs.get("years_input", []) if scen_name == "Base" else [],
            }
            scenarios.append(scen)

        commodities.append({
            "commodity":     comm,
            "is_primary":    ci == 0,
            "has_scenarios": True,
            "scenarios":     scenarios,
        })

    return {
        "source_file": os.path.basename(wb_path),
        "mine": {
            "mine_name":         "Mine C3",
            "license_number":    "12891L",
            "country":           "Mozambique",
            "province":          "Zambezia",
            "headline":          "CHINAKA RESOURCE MINING 3",
            "subtitle":          "REE + Monazite + Spodumene",
            "ore_reserve":       50,
            "reserve_unit":      "Mt",
            "throughput_pa":     2.5,
            "throughput_unit":   "Mtpa",
            "life_of_mine_yr":   20,
            "wacc":              0.10,
            "tax_rate":          0.30,
            "royalty_rate":      0.05,
            "ramp_up_y1":        0.40,
            "ramp_up_y2":        0.70,
            "ramp_up_y3":        1.00,
            "closure_rehab_cost": 120e6,
            "source_file":       os.path.basename(wb_path),
        },
        "commodities": commodities,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# MINE11  (M'Gomo — Gold + Graphite Bear/Base/Bull)
# ═══════════════════════════════════════════════════════════════════════════════

def _parse_mine11_sheet(ws) -> dict:
    yc = _year_cols(ws)
    prod_d  = _read_input_row(ws, 5,  yc)
    price_d = _read_input_row(ws, 8,  yc)
    opex_d  = _read_input_row(ws, 14, yc)
    dep_d   = _read_input_row(ws, 19, yc)
    capex_d = _read_input_row(ws, 24, yc)

    prod_sc   = _derive_scalar(prod_d,  is_prod=True)
    price_sc  = _derive_scalar(price_d, is_price=True)
    capex_sc  = _derive_scalar(capex_d, is_capex=True)
    opex_sc   = _derive_scalar(opex_d,  is_opex=True)
    dep_pa    = _derive_dep(dep_d)

    years_input = []
    for yr in sorted(yc.keys()):
        years_input.append({
            "year":            yr,
            "production":      prod_d.get(yr),
            "commodity_price": price_d.get(yr),
            "operating_costs": opex_d.get(yr),
            "depreciation":    dep_d.get(yr),
            "capex":           capex_d.get(yr),
        })

    # Try to read metrics from rows 32-37
    metrics_col = 4
    for tc in [4, 3, 2]:
        v = _cell(ws, 32, tc)
        if v: metrics_col = tc; break

    def _met_val(r):
        v = _cell(ws, r, metrics_col)
        if v is None: return None
        s = str(v).replace("$","").replace("mm","").replace("%","").replace(",","").strip()
        try: return float(s)
        except: return None

    metrics = {
        "npv": _met_val(32), "irr": _met_val(33),
        "payback": _s(_cell(ws, 34, metrics_col)),
        "moic": _met_val(35),
        "total_capex": _met_val(36), "total_lom_fcf": _met_val(37),
    }

    return {**prod_sc, **price_sc, **capex_sc, **opex_sc,
            "depreciation_pa": dep_pa,
            "years_input": years_input,
            "metrics": metrics}


def parse_mine11(wb_path: str) -> dict:
    wb = openpyxl.load_workbook(wb_path, data_only=True)

    # ── Gold (Single) ─────────────────────────────────────────────────────────
    gold = _parse_mine11_sheet(wb["Gold"])
    # avg_recovered_grade for Gold: 0.24 g/m3 (from Mine_Profile notes)
    # annual_production = throughput_pa × grade / 1000 = 636480 × 0.24 / 1000 = 152.755 kg/yr
    gold_scenario = {
        "scenario":              "Single",
        "sheet_name":            "Gold",
        "price_base":            gold.get("price_base"),
        "price_unit":            "$/kg",
        "price_escalation_rate": 0.0,  # Gold price is fixed (Mine Profile: $48,000/kg)
        "annual_production":     gold.get("annual_production"),
        "avg_recovered_grade":   0.24,   # g/m3 — used by run_dcf to derive production
        "production_unit":       "kg",
        "opex_steady_state":     gold.get("opex_steady_state"),
        "opex_escalation_rate":  gold.get("opex_escalation_rate", 0.0),
        "initial_capex":         gold.get("initial_capex"),
        "sustaining_capex_pa":   gold.get("sustaining_capex_pa", 0),
        "capex_deployment_year": gold.get("capex_deployment_year", 0),
        "depreciation_pa":       gold.get("depreciation_pa"),
        "production_start_year": gold.get("production_start_year", 1),
        "royalty_rate":          0.06,
        "basis_notes":           "Alluvial gold — single scenario",
        "years_input":           gold.get("years_input", []),
        "ingested_metrics":      gold.get("metrics"),
    }

    # ── Graphite Bear / Base / Bull ───────────────────────────────────────────
    graphite_scenarios = []
    for scen_name in ["Bear", "Base", "Bull"]:
        sname = f"Graphite {scen_name}"
        if sname not in wb.sheetnames:
            continue
        g = _parse_mine11_sheet(wb[sname])
        graphite_scenarios.append({
            "scenario":              scen_name,
            "sheet_name":            sname,
            "price_base":            g.get("price_base"),
            "price_unit":            "$/t",
            "price_escalation_rate": g.get("price_escalation_rate", 0.0),
            "annual_production":     g.get("annual_production"),
            "production_unit":       "t conc",
            "opex_steady_state":     g.get("opex_steady_state"),
            "opex_escalation_rate":  g.get("opex_escalation_rate", 0.0),
            "initial_capex":         g.get("initial_capex"),
            "sustaining_capex_pa":   g.get("sustaining_capex_pa", 0),
            "capex_deployment_year": g.get("capex_deployment_year", 2),
            "depreciation_pa":       g.get("depreciation_pa"),
            "production_start_year": g.get("production_start_year", 3),
            "royalty_rate":          0.03,
            "basis_notes":           f"Graphite {scen_name} case",
            "years_input":           g.get("years_input", []),
            "ingested_metrics":      g.get("metrics"),
        })

    return {
        "source_file": os.path.basename(wb_path),
        "mine": {
            "mine_name":         "Mine G",
            "license_number":    "9015L",
            "country":           "Mozambique",
            "province":          "Tete",
            "headline":          "M'GOMO MINE",
            "subtitle":          "Gold + Graphite",
            "ore_reserve":       31891000,   # 31.891 Mm3 → m3 (same unit as throughput_pa)
            "reserve_unit":      "m3",
            "throughput_pa":     636480,
            "throughput_unit":   "m3/yr",
            "life_of_mine_yr":   50,
            "wacc":              0.15,
            "tax_rate":          0.32,
            "royalty_rate":      0.06,
            "ramp_up_y1":        0.35,
            "ramp_up_y2":        0.70,
            "ramp_up_y3":        1.00,
            "closure_rehab_cost": 0,
            "source_file":       os.path.basename(wb_path),
        },
        "commodities": [
            {
                "commodity":     "Gold",
                "is_primary":    True,
                "has_scenarios": False,
                "scenarios":     [gold_scenario],
            },
            {
                "commodity":     "Graphite",
                "is_primary":    False,
                "has_scenarios": True,
                "scenarios":     graphite_scenarios,
            },
        ],
    }


# ═══════════════════════════════════════════════════════════════════════════════
# SUPABASE WRITER
# ═══════════════════════════════════════════════════════════════════════════════

def write_payload(payload: dict, dry_run: bool = False) -> Optional[str]:
    mine_data = {k: v for k, v in payload["mine"].items() if v is not None}
    print(f"\n── {mine_data.get('mine_name')} ({mine_data.get('license_number')}) ──")

    if dry_run:
        print(json.dumps(payload, indent=2, default=str))
        return None

    client = _sb()

    # Upsert mine
    lic = mine_data.get("license_number")
    existing = client.table("m4_mines").select("id").eq("license_number", lic).execute() if lic else None
    if existing and existing.data:
        mine_id = existing.data[0]["id"]
        client.table("m4_mines").update(mine_data).eq("id", mine_id).execute()
        print(f"  Mine updated: {mine_id}")
    else:
        res = client.table("m4_mines").insert(mine_data).execute()
        mine_id = res.data[0]["id"]
        print(f"  Mine created: {mine_id}")

    # Upsert commodities + scenarios
    for ci, comm_data in enumerate(payload["commodities"]):
        comm_payload = {
            "mine_id":       mine_id,
            "commodity":     comm_data["commodity"],
            "is_primary":    comm_data.get("is_primary", False),
            "has_scenarios": comm_data.get("has_scenarios", False),
            "display_order": ci,
        }
        existing_c = (client.table("m4_commodities").select("id")
                     .eq("mine_id", mine_id).eq("commodity", comm_data["commodity"]).execute())
        if existing_c.data:
            comm_id = existing_c.data[0]["id"]
            client.table("m4_commodities").update(comm_payload).eq("id", comm_id).execute()
        else:
            res = client.table("m4_commodities").insert(comm_payload).execute()
            comm_id = res.data[0]["id"]
        print(f"  Commodity '{comm_data['commodity']}' → {comm_id}")

        for scen_data in comm_data.get("scenarios", []):
            years_input = scen_data.pop("years_input", [])
            ingested_metrics = scen_data.pop("ingested_metrics", None)

            scen_payload = {k: v for k, v in {
                "commodity_id":           comm_id,
                "scenario":               scen_data.get("scenario", "Single"),
                "sheet_name":             scen_data.get("sheet_name"),
                "price_base":             scen_data.get("price_base"),
                "price_unit":             scen_data.get("price_unit", "$/t"),
                "price_escalation_rate":  scen_data.get("price_escalation_rate"),
                "annual_production":      scen_data.get("annual_production"),
                "production_unit":        scen_data.get("production_unit", "tonnes"),
                "opex_steady_state":      scen_data.get("opex_steady_state"),
                "opex_escalation_rate":   scen_data.get("opex_escalation_rate"),
                "initial_capex":          scen_data.get("initial_capex"),
                "sustaining_capex_pa":    scen_data.get("sustaining_capex_pa"),
                "capex_deployment_year":  scen_data.get("capex_deployment_year"),
                "depreciation_pa":        scen_data.get("depreciation_pa"),
                "avg_depreciation_years": scen_data.get("avg_depreciation_years"),
                "avg_recovered_grade":    scen_data.get("avg_recovered_grade"),
                "production_start_year":  scen_data.get("production_start_year"),
                "wacc":                   scen_data.get("wacc"),
                "royalty_rate":           scen_data.get("royalty_rate"),
                "basis_notes":            scen_data.get("basis_notes"),
            }.items() if v is not None}

            existing_s = (client.table("m4_scenarios").select("id")
                         .eq("commodity_id", comm_id).eq("scenario", scen_payload["scenario"]).execute())
            if existing_s.data:
                scen_id = existing_s.data[0]["id"]
                client.table("m4_scenarios").update(scen_payload).eq("id", scen_id).execute()
            else:
                res = client.table("m4_scenarios").insert(scen_payload).execute()
                scen_id = res.data[0]["id"]
            print(f"    Scenario '{scen_payload['scenario']}': {scen_id}")

            # Write input years (only non-None rows)
            if years_input:
                client.table("m4_dcf_inputs").delete().eq("scenario_id", scen_id).execute()
                rows = []
                for y in years_input:
                    r = {"scenario_id": scen_id, "year": y["year"]}
                    for f in ["production","commodity_price","operating_costs","depreciation","capex"]:
                        if y.get(f) is not None: r[f] = y[f]
                    rows.append(r)
                client.table("m4_dcf_inputs").insert(rows).execute()
                print(f"    → {len(rows)} input year rows written")

            # Metrics are NOT ingested from Excel — run_dcf is the sole source of truth.
            # m4_metrics is populated by the auto-calculate in list_mines / get_scenarios.

    return mine_id


# ═══════════════════════════════════════════════════════════════════════════════
# CLI
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="Ingest workbooks into m4_* tables")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--json", metavar="FILE")
    args = parser.parse_args()

    workbooks = [
        (BASE_DIR / "Mining Investment Model Prototype (4.20.2026).xlsx", parse_prototype),
        (BASE_DIR / "Mine1.xlsx",   parse_mine1),
        (BASE_DIR / "mine11.xlsx",  parse_mine11),
    ]

    all_payloads = []
    for wb_path, parser_fn in workbooks:
        print(f"\nParsing: {wb_path.name}")
        try:
            payload = parser_fn(str(wb_path))
            all_payloads.append(payload)
            write_payload(payload, dry_run=args.dry_run)
        except Exception as e:
            print(f"  ERROR: {e}", file=sys.stderr)
            import traceback; traceback.print_exc()

    if args.json:
        with open(args.json, "w") as f:
            json.dump(all_payloads, f, indent=2, default=str)
        print(f"\nJSON written to {args.json}")


if __name__ == "__main__":
    main()
