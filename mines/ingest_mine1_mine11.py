"""
ingest_mine1_mine11.py
======================
Ingest Mine1.xlsx and mine11.xlsx into the schema_v2 / schema_v3 tables.

WHAT IT READS:
  For every DCF sheet it reads ONLY the input rows:
    Row  5  — production (t/yr or kg/yr)
    Row  8  — commodity price (base)
    Row 14  — operating costs (OPEX, p.a.)
    Row 19  — depreciation (p.a.)
    Row 24  — CAPEX (initial; year-0 column)
  Plus the summary metrics block (rows 32–37) for ingested scenario_metrics.

  All other rows (9–11, 15–16, 20–21, 25–28) are calculated by the DCF engine
  and are NOT stored.

FILE STRUCTURE ASSUMED:
  Mine1.xlsx    → sheets: "REE", "Monazite", "Spodumene"   (each single Base DCF)
  mine11.xlsx   → sheets: "Gold", "Graphite Bear", "Graphite Base", "Graphite Bull"

COMMODITY DETECTION:
  Sheet names are mapped to commodities automatically.
  Bear/Base/Bull suffix → scenario enum.

USAGE:
  pip install openpyxl supabase
  export SUPABASE_URL=...
  export SUPABASE_SERVICE_KEY=...

  python ingest_mine1_mine11.py --dry-run Mine1.xlsx mine11.xlsx
  python ingest_mine1_mine11.py Mine1.xlsx mine11.xlsx
  python ingest_mine1_mine11.py --json out.json Mine1.xlsx mine11.xlsx
"""
from __future__ import annotations
import argparse, json, os, re, sys
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.utils import get_column_letter

# ── Supabase ──────────────────────────────────────────────────────────────────
SUPABASE_URL = os.environ.get("SUPABASE_URL",
    "https://snbnqwrxvptrfjsecljd.supabase.co")
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_KEY",
    os.environ.get("SUPABASE_ANON_KEY",
    "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9."
    "eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InNuYm5xd3J4dnB0cmZqc2VjbGpkIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3OTg5MzQ4MSwiZXhwIjoyMDk1NDY5NDgxfQ."
    "t92K9HW0uQpCWy08c6CPtGKynHN5ET3ymcfcupNJTO0"))


def _sb():
    from supabase import create_client
    return create_client(SUPABASE_URL, SUPABASE_KEY)


# ── DCF input row numbers (1-indexed Excel rows) ──────────────────────────────
INPUT_ROWS = {
    "production":      5,
    "commodity_price": 8,
    "operating_costs": 14,
    "depreciation":    19,
    "capex":           24,
}

METRICS_ROWS = {
    "npv":           32,
    "irr":           33,
    "payback":       34,
    "moic":          35,
    "total_capex":   36,
    "total_lom_fcf": 37,
}

# ── Sheet → (commodity, scenario) mapping ─────────────────────────────────────
SCENARIO_ENUM = {"Bear", "Base", "Bull", "Single"}

def _detect_commodity_scenario(sheet_name: str) -> Tuple[str, str]:
    """
    'Graphite Bear' → ('Graphite', 'Bear')
    'REE'           → ('REE', 'Single')
    'Gold'          → ('Gold', 'Single')
    """
    parts = sheet_name.strip().split()
    if len(parts) >= 2 and parts[-1] in SCENARIO_ENUM:
        commodity = " ".join(parts[:-1])
        scenario  = parts[-1]
    else:
        commodity = sheet_name.strip()
        scenario  = "Single"
    return commodity, scenario


# ── Workbook metadata extraction ──────────────────────────────────────────────

def _find_header_sheet(wb: openpyxl.Workbook) -> Optional[openpyxl.worksheet.worksheet.Worksheet]:
    """Return the Exec Summary or first non-DCF sheet for mine identity."""
    for name in ["Exec Summary", "Executive Summary", "Summary", "Mine Profile"]:
        if name in wb.sheetnames:
            return wb[name]
    return None


def _cell(ws, row: int, col: int):
    """Return numeric value or string from cell; None if empty/formula."""
    v = ws.cell(row=row, column=col).value
    if v is None or str(v).startswith("="):
        return None
    return v


def _num(v) -> Optional[float]:
    try:
        return float(v) if v is not None else None
    except (ValueError, TypeError):
        return None


def _int(v) -> Optional[int]:
    f = _num(v)
    return int(round(f)) if f is not None else None


# ── DCF sheet parser ──────────────────────────────────────────────────────────

def parse_dcf_sheet(ws: openpyxl.worksheet.worksheet.Worksheet,
                    sheet_name: str) -> Dict:
    """
    Parse one DCF sheet.  Returns:
    {
      sheet_name, commodity, scenario,
      wacc, price_base, price_unit, price_escalation_rate,
      opex, opex_escalation_rate, initial_capex, depreciation_pa,
      sustaining_capex_pa, production_start_year,
      annual_production,
      years: [{year, production, commodity_price, operating_costs,
               depreciation, capex}],
      metrics: {npv, irr, payback, moic, total_capex, total_lom_fcf}
    }
    """
    commodity, scenario = _detect_commodity_scenario(sheet_name)

    # Find year columns: row 4 is usually the year header (Year 0, Year 1 …)
    # Columns start at B (col 2) for Year 0
    year_cols: Dict[int, int] = {}   # year_number → column_index
    max_col = ws.max_column or 50
    for col in range(2, min(max_col + 1, 60)):
        hdr = ws.cell(row=4, column=col).value
        if hdr is None:
            continue
        s = str(hdr).strip().lower().replace(" ", "").replace("year", "").replace("yr", "")
        try:
            year_num = int(float(s))
            year_cols[year_num] = col
        except ValueError:
            pass

    # Fallback: treat col 2..N as Year 0..N-2 if no header found
    if not year_cols:
        max_data_col = min(ws.max_column or 30, 40)
        for i, col in enumerate(range(2, max_data_col + 1)):
            year_cols[i] = col

    # Extract WACC from top of sheet (common positions: B2 or B3)
    wacc = None
    for r in [2, 3]:
        for c in [2, 3]:
            v = _num(_cell(ws, r, c))
            if v is not None and 0.05 <= v <= 0.40:
                wacc = v
                break
        if wacc:
            break

    # Extract price unit from row 8 label in col A
    price_unit_raw = ws.cell(row=INPUT_ROWS["commodity_price"], column=1).value or ""
    price_unit = str(price_unit_raw).strip()[:20] or "$/t"

    # --- Pull input rows across all year columns ---
    years_data: List[Dict] = []
    prod_row   = INPUT_ROWS["production"]
    price_row  = INPUT_ROWS["commodity_price"]
    opex_row   = INPUT_ROWS["operating_costs"]
    dep_row    = INPUT_ROWS["depreciation"]
    capex_row  = INPUT_ROWS["capex"]

    for year_num in sorted(year_cols):
        col = year_cols[year_num]
        years_data.append({
            "year":            year_num,
            "production":      _num(_cell(ws, prod_row,  col)),
            "commodity_price": _num(_cell(ws, price_row, col)),
            "operating_costs": _num(_cell(ws, opex_row,  col)),
            "depreciation":    _num(_cell(ws, dep_row,   col)),
            "capex":           _num(_cell(ws, capex_row, col)),
        })

    # --- Derive scenario-level inputs from year data ---
    # Price base: first non-zero price across all years
    price_base = next((r["commodity_price"] for r in years_data
                       if r["commodity_price"] and r["commodity_price"] > 0), None)

    # Price escalation: (last_price / first_price)^(1/(n-1)) - 1
    prices = [r["commodity_price"] for r in years_data
              if r["commodity_price"] and r["commodity_price"] > 0]
    price_escalation_rate = None
    if len(prices) >= 2:
        try:
            n = len(prices) - 1
            price_escalation_rate = round((prices[-1] / prices[0]) ** (1 / n) - 1, 4)
            if abs(price_escalation_rate) > 0.20:     # cap wild outliers
                price_escalation_rate = None
        except (ZeroDivisionError, ValueError):
            pass

    # OPEX at steady state: median of non-zero positive opex years
    opex_vals = sorted(r["operating_costs"] for r in years_data
                       if r["operating_costs"] and r["operating_costs"] > 0)
    opex_ss = opex_vals[len(opex_vals) // 2] if opex_vals else None

    # OPEX escalation: (last / first)^(1/(n-1))-1
    opex_esc = None
    if len(opex_vals) >= 2:
        try:
            n = len(opex_vals) - 1
            opex_esc = round((opex_vals[-1] / opex_vals[0]) ** (1 / n) - 1, 4)
            if abs(opex_esc) > 0.20:
                opex_esc = None
        except (ZeroDivisionError, ValueError):
            pass

    # Initial CAPEX: largest negative capex value (year 0 or year 1)
    initial_capex = None
    for r in years_data[:3]:   # usually year 0 or 1
        c = r["capex"]
        if c is not None and c < 0:
            v = abs(c)
            if initial_capex is None or v > initial_capex:
                initial_capex = v

    # Sustaining CAPEX: median of non-zero post-construction capex
    sust_vals = sorted(abs(r["capex"]) for r in years_data[3:]
                       if r["capex"] and abs(r["capex"]) > 0)
    sustaining_capex_pa = sust_vals[len(sust_vals) // 2] if sust_vals else 0.0

    # Depreciation p.a.: median of non-zero depreciation years
    dep_vals = sorted(r["depreciation"] for r in years_data
                      if r["depreciation"] and r["depreciation"] > 0)
    depreciation_pa = dep_vals[len(dep_vals) // 2] if dep_vals else None

    # Annual production: max production across years (steady-state peak)
    annual_production = max((r["production"] or 0) for r in years_data) or None

    # Production start year: first year with non-zero production
    production_start_year = next((r["year"] for r in years_data
                                  if r["production"] and r["production"] > 0), 1)

    # CAPEX deployment year: first year with negative capex
    capex_deployment_year = next((r["year"] for r in years_data
                                  if r["capex"] and r["capex"] < 0), 0)

    # --- Metrics block (rows 32–37) ---
    metrics = {}
    # Metrics are in summary column — usually the last data column or a fixed col
    met_col = max(year_cols.values()) + 1 if year_cols else 2
    # Try fixed column D (col 4) first, then last col+1
    for try_col in [4, 3, met_col]:
        npv_v = _num(_cell(ws, METRICS_ROWS["npv"], try_col))
        if npv_v is not None:
            met_col = try_col
            break

    metrics["npv"]           = _num(_cell(ws, METRICS_ROWS["npv"],           met_col))
    metrics["irr"]           = _num(_cell(ws, METRICS_ROWS["irr"],           met_col))
    metrics["payback"]       = _cell(ws, METRICS_ROWS["payback"],             met_col)
    metrics["moic"]          = _num(_cell(ws, METRICS_ROWS["moic"],          met_col))
    metrics["total_capex"]   = _num(_cell(ws, METRICS_ROWS["total_capex"],   met_col))
    metrics["total_lom_fcf"] = _num(_cell(ws, METRICS_ROWS["total_lom_fcf"], met_col))

    # Normalise payback to string
    if metrics["payback"] is not None:
        pv = metrics["payback"]
        if isinstance(pv, (int, float)):
            metrics["payback"] = f"{int(pv)} year(s)"
        else:
            metrics["payback"] = str(pv)

    return {
        "sheet_name":              sheet_name,
        "commodity":               commodity,
        "scenario":                scenario,
        "wacc":                    wacc,
        "price_base":              price_base,
        "price_unit":              price_unit,
        "price_escalation_rate":   price_escalation_rate,
        "opex":                    opex_ss,
        "opex_escalation_rate":    opex_esc,
        "initial_capex":           initial_capex,
        "sustaining_capex_pa":     sustaining_capex_pa,
        "depreciation_pa":         depreciation_pa,
        "annual_production":       annual_production,
        "production_start_year":   production_start_year,
        "capex_deployment_year":   capex_deployment_year,
        "years":                   years_data,
        "metrics":                 metrics,
    }


# ── Workbook-level parser ─────────────────────────────────────────────────────

# Sheet names to skip (not DCF sheets)
_SKIP_SHEETS = {
    "exec summary", "executive summary", "summary", "mine profile",
    "ree scenario analysis", "scenario analysis", "assumptions",
    "cover", "contents", "index", "notes", "instructions",
}

# Known commodity sheets
_COMMODITY_SHEETS = {
    "ree", "monazite", "spodumene", "gold",
    "graphite", "graphite bear", "graphite base", "graphite bull",
    "combined",
}

def _is_dcf_sheet(name: str) -> bool:
    nl = name.strip().lower()
    if nl in _SKIP_SHEETS:
        return False
    # Heuristic: name contains a known commodity OR ends with Bear/Base/Bull/Single
    for c in _COMMODITY_SHEETS:
        if nl == c or nl.startswith(c + " "):
            return True
    return nl.endswith(("bear", "base", "bull", "single"))


def parse_workbook(wb_path: str) -> Dict:
    """
    Parse all DCF sheets in a workbook.
    Returns {
      source_file,
      mine_name,   license_number,
      dcf_sheets: [parsed_sheet_dict, ...]
    }
    """
    wb = openpyxl.load_workbook(wb_path, data_only=True)

    # Try to get mine identity from Exec Summary sheet
    mine_name      = os.path.splitext(os.path.basename(wb_path))[0]
    license_number = re.sub(r"[^0-9A-Z]", "", mine_name.upper())[:12]

    hdr = _find_header_sheet(wb)
    if hdr:
        # Try row 1 col B or col A for mine name
        for r in [1, 2]:
            v = hdr.cell(row=r, column=2).value or hdr.cell(row=r, column=1).value
            if v and len(str(v).strip()) > 2:
                mine_name = str(v).strip()
                break
        # License: look for pattern like 12891L or 9015L in first 10 rows
        for r in range(1, 10):
            for c in range(1, 5):
                v = str(hdr.cell(row=r, column=c).value or "")
                m = re.search(r"\b(\d{4,6}[A-Z])\b", v)
                if m:
                    license_number = m.group(1)
                    break

    dcf_sheets = []
    for sn in wb.sheetnames:
        if not _is_dcf_sheet(sn):
            continue
        try:
            parsed = parse_dcf_sheet(wb[sn], sn)
            dcf_sheets.append(parsed)
        except Exception as e:
            print(f"  WARNING: skipped sheet '{sn}': {e}", file=sys.stderr)

    return {
        "source_file":    os.path.basename(wb_path),
        "mine_name":      mine_name,
        "license_number": license_number,
        "dcf_sheets":     dcf_sheets,
    }


# ── Supabase writer ───────────────────────────────────────────────────────────

def write_to_supabase(payload: Dict, dry_run: bool = False) -> Optional[str]:
    """
    Write parsed workbook payload to schema_v2 tables.
    Returns mine_id on success.
    """
    client = None if dry_run else _sb()
    mine_name      = payload["mine_name"]
    license_number = payload["license_number"]
    source_file    = payload["source_file"]
    dcf_sheets     = payload["dcf_sheets"]

    if not dcf_sheets:
        print(f"  No DCF sheets found in {source_file}")
        return None

    # Derive unique commodities
    commodities: Dict[str, Dict] = {}   # commodity_name → {has_scenarios, is_primary}
    for s in dcf_sheets:
        c = s["commodity"]
        if c not in commodities:
            commodities[c] = {"has_scenarios": False, "is_primary": False}
        if s["scenario"] in ("Bear", "Bull"):
            commodities[c]["has_scenarios"] = True

    # Mark first commodity as primary
    first_comm = next(iter(commodities))
    commodities[first_comm]["is_primary"] = True

    print(f"\n── {mine_name} ({license_number}) ──")
    print(f"   Commodities: {list(commodities)}")
    print(f"   Sheets parsed: {[s['sheet_name'] for s in dcf_sheets]}")

    if dry_run:
        print(json.dumps(payload, indent=2, default=str))
        return None

    # ── Upsert mine ───────────────────────────────────────────────
    existing_mine = client.table("mines").select("id").eq("license_number", license_number).execute()
    if existing_mine.data:
        mine_id = existing_mine.data[0]["id"]
        print(f"   Mine exists: {mine_id} — updating")
    else:
        res = client.table("mines").insert({
            "mine_name":      mine_name,
            "license_number": license_number,
            "source_file":    source_file,
        }).execute()
        mine_id = res.data[0]["id"]
        print(f"   Mine created: {mine_id}")

    # ── Upsert commodities ────────────────────────────────────────
    comm_id_map: Dict[str, str] = {}    # commodity_name → commodity_id
    for comm_name, comm_meta in commodities.items():
        existing_comm = (client.table("mine_commodities").select("id")
                        .eq("mine_id", mine_id).eq("commodity", comm_name).execute())
        if existing_comm.data:
            comm_id = existing_comm.data[0]["id"]
        else:
            res = client.table("mine_commodities").insert({
                "mine_id":      mine_id,
                "commodity":    comm_name,
                "is_primary":   comm_meta["is_primary"],
                "has_scenarios": comm_meta["has_scenarios"],
            }).execute()
            comm_id = res.data[0]["id"]
        comm_id_map[comm_name] = comm_id
        print(f"   Commodity '{comm_name}' → {comm_id}")

    # ── Upsert commodity_scenarios + dcf_years (input rows only) ─
    for s in dcf_sheets:
        comm_id = comm_id_map[s["commodity"]]
        scen_enum = s["scenario"]

        # Upsert scenario row
        existing_scen = (client.table("commodity_scenarios").select("id")
                        .eq("commodity_id", comm_id).eq("scenario", scen_enum).execute())
        scen_payload = {
            "commodity_id":          comm_id,
            "scenario":              scen_enum,
            "sheet_name":            s["sheet_name"],
            "wacc":                  s["wacc"],
            "price_base":            s["price_base"],
            "price_unit":            s["price_unit"],
            "price_escalation_rate": s["price_escalation_rate"],
            "opex":                  s["opex"],
            "opex_escalation_rate":  s["opex_escalation_rate"],
            "initial_capex":         s["initial_capex"],
            "sustaining_capex_pa":   s["sustaining_capex_pa"],
            "depreciation_pa":       s["depreciation_pa"],
            "annual_production":     s["annual_production"],
            "production_start_year": s["production_start_year"],
            "capex_deployment_year": s["capex_deployment_year"],
        }
        # Remove None values
        scen_payload = {k: v for k, v in scen_payload.items() if v is not None}

        if existing_scen.data:
            scen_id = existing_scen.data[0]["id"]
            client.table("commodity_scenarios").update(scen_payload).eq("id", scen_id).execute()
            print(f"   Scenario '{scen_enum}' updated: {scen_id}")
        else:
            res = client.table("commodity_scenarios").insert(scen_payload).execute()
            scen_id = res.data[0]["id"]
            print(f"   Scenario '{scen_enum}' created: {scen_id}")

        # Write dcf_years (input rows, source='ingested')
        # Delete existing ingested rows first to avoid duplicates
        client.table("dcf_years").delete().eq("scenario_id", scen_id).eq("source", "ingested").execute()

        year_rows = []
        for y in s["years"]:
            row = {
                "scenario_id":     scen_id,
                "source":          "ingested",
                "year":            y["year"],
                "production":      y["production"],
                "commodity_price": y["commodity_price"],
                "operating_costs": y["operating_costs"],
                "depreciation":    y["depreciation"],
                "capex":           y["capex"],
            }
            year_rows.append({k: v for k, v in row.items() if v is not None or k in ("year", "scenario_id", "source")})

        if year_rows:
            client.table("dcf_years").insert(year_rows).execute()
            print(f"   Inserted {len(year_rows)} DCF year rows for '{scen_enum}'")

        # Write scenario_metrics
        m = s["metrics"]
        if any(v is not None for v in m.values()):
            existing_met = client.table("scenario_metrics").select("scenario_id").eq("scenario_id", scen_id).execute()
            met_payload  = {k: v for k, v in m.items() if v is not None}
            met_payload["scenario_id"] = scen_id
            if existing_met.data:
                client.table("scenario_metrics").update(met_payload).eq("scenario_id", scen_id).execute()
            else:
                client.table("scenario_metrics").insert(met_payload).execute()
            print(f"   Metrics written for '{scen_enum}'")

    return mine_id


# ── CLI entry point ───────────────────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Ingest Mine1/mine11 Excel workbooks into schema_v2")
    parser.add_argument("workbooks", nargs="+", help="Excel file paths")
    parser.add_argument("--dry-run", action="store_true", help="Parse only, no DB writes")
    parser.add_argument("--json",    metavar="FILE",       help="Dump parsed payload to JSON file")
    args = parser.parse_args()

    all_payloads = []
    for wb_path in args.workbooks:
        print(f"\nParsing: {wb_path}")
        try:
            payload = parse_workbook(wb_path)
            all_payloads.append(payload)
            if not args.dry_run:
                write_to_supabase(payload)
            else:
                write_to_supabase(payload, dry_run=True)
        except Exception as e:
            print(f"  ERROR: {e}", file=sys.stderr)
            if "--verbose" in sys.argv:
                import traceback; traceback.print_exc()

    if args.json:
        with open(args.json, "w") as f:
            json.dump(all_payloads, f, indent=2, default=str)
        print(f"\nJSON written to {args.json}")
