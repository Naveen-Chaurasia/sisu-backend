"""
ingest_benchmarked.py
─────────────────────
Ingest the two benchmarked mining workbooks into the Supabase schema (schema.sql).

  • Mine1_12891L_Revised_Benchmarked_Apr2026.xlsx
      Commodities: REE, Monazite, Spodumene, Combined DCF
      Scenarios:   on a separate 'REE Scenario Analysis' sheet (kept as exec rows)

  • Complex_Release_2_Mine11_9015L_..._Graphite_3_scenarios_Apr2026.xlsx
      Commodities: Gold (single), Graphite (Bear/Base/Bull as 3 DCF sheets)

Both share an identical DCF row map, so ONE parser handles both. Differences
(WACC, year count, scenario style) are detected from the sheets themselves.

USAGE
  pip install openpyxl supabase
  export SUPABASE_URL=...            # https://xxxx.supabase.co
  export SUPABASE_SERVICE_KEY=...    # service_role key (server side only)

  # Dry run — parse + print, write nothing:
  python ingest_benchmarked.py --dry-run mine1.xlsx mine11.xlsx

  # Real ingest:
  python ingest_benchmarked.py mine1.xlsx mine11.xlsx

  # Dump parsed payload to JSON (no DB needed):
  python ingest_benchmarked.py --json out.json mine1.xlsx mine11.xlsx
"""
from __future__ import annotations
import argparse, json, os, re, sys
from typing import Any, Optional
import openpyxl


# ════════════════════════════════════════════════════════════════════
# Fixed DCF row map — identical across every commodity sheet in both files
# ════════════════════════════════════════════════════════════════════
DCF_ROWS = {
    "production":      5,
    "commodity_price": 8,
    "gross_revenue":   9,
    "royalty":         10,
    "net_revenue":     11,
    "operating_costs": 14,
    "ebitda":          15,
    "ebitda_margin":   16,
    "depreciation":    19,
    "ebit":            20,
    "income_tax":      21,
    "capex":           24,
    "free_cash_flow":  25,
    "cumulative_fcf":  26,
    "discount_factor": 27,
    "discounted_cf":   28,
}
METRIC_ROWS = {            # the R32–R37 summary block (value in col C)
    "npv": 32, "irr": 33, "payback": 34,
    "moic": 35, "total_capex": 36, "total_lom_fcf": 37,
}
HEADER_ROW = 3             # 'LINE ITEM | Unit | Yr 0 | Yr 1 | ...'
FIRST_YEAR_COL = 3         # column C = Yr 0


# ════════════════════════════════════════════════════════════════════
# Parsing helpers — tolerant of '$826mm', '16.6%', '7 year(s)', floats
# ════════════════════════════════════════════════════════════════════
def to_num(v: Any) -> Optional[float]:
    """Extract a float from messy money/percent/multiple strings."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s or s.lower() in {"n/a", "—", "-", "beyond lom", "no payback in lom"}:
        return None
    s = s.replace(",", "").replace("$", "").replace("mm", "").replace("x", "").replace("%", "")
    m = re.search(r"-?\d+\.?\d*", s)
    return float(m.group()) if m else None


def pct_to_frac(v: Any) -> Optional[float]:
    """'16.6%' -> 0.166 ; a bare 0.572 (already fraction) is left as-is."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)            # EBITDA margin cells are already fractions
    n = to_num(v)
    return n / 100.0 if n is not None else None


def clean(v: Any) -> Optional[str]:
    if v is None:
        return None
    s = str(v).replace("\n", " ").strip()
    return s or None


def split_banner(text: Optional[str]) -> list[str]:
    return [p.strip() for p in (text or "").split("|") if p.strip()]


def find_wacc(ws) -> Optional[float]:
    """Pull '15% WACC' / '10% WACC' from the row-1/row-2 banners or R27 label."""
    for r in (1, 2, 27):
        cell = clean(ws.cell(row=r, column=1).value)
        if cell:
            m = re.search(r"(\d+(?:\.\d+)?)\s*%\s*WACC", cell)
            if m:
                return float(m.group(1)) / 100.0
    return None


# ════════════════════════════════════════════════════════════════════
# Core: parse one commodity DCF sheet into a dict
# ════════════════════════════════════════════════════════════════════
def parse_dcf_sheet(ws) -> dict:
    # how many year columns are populated on the header row?
    years: list[int] = []
    col = FIRST_YEAR_COL
    while True:
        h = ws.cell(row=HEADER_ROW, column=col).value
        if h is None or not str(h).strip().lower().startswith("yr"):
            break
        years.append(int(re.search(r"\d+", str(h)).group()))
        col += 1
    last_col = FIRST_YEAR_COL + len(years) - 1

    # year-by-year rows
    dcf = []
    for idx, yr in enumerate(years):
        c = FIRST_YEAR_COL + idx
        rec = {"year": yr}
        for field, row in DCF_ROWS.items():
            raw = ws.cell(row=row, column=c).value
            rec[field] = pct_to_frac(raw) if field == "ebitda_margin" else to_num(raw)
        dcf.append(rec)

    # summary metric block (value lives in col C)
    metrics: dict[str, Any] = {}
    for key, row in METRIC_ROWS.items():
        raw = ws.cell(row=row, column=FIRST_YEAR_COL).value
        if key == "payback":
            metrics[key] = clean(raw)
        elif key == "irr":
            metrics[key] = pct_to_frac(raw)
        else:
            metrics[key] = to_num(raw)

    banner1 = split_banner(clean(ws.cell(row=1, column=1).value))
    banner2 = split_banner(clean(ws.cell(row=2, column=1).value))

    # price + unit from R8 (first producing year is most reliable: use Yr 0/1)
    price_unit = clean(ws.cell(row=DCF_ROWS["commodity_price"], column=2).value)
    price_base = next((d["commodity_price"] for d in dcf if d["commodity_price"]), None)

    # benchmark basis note (Mine11 graphite puts it on R39; else row 2 tail)
    basis = None
    for r in range(38, 42):
        t = clean(ws.cell(row=r, column=1).value)
        if t and "BENCHMARK" in t.upper():
            basis = t
            break
    if not basis and len(banner2) > 1:
        basis = " | ".join(banner2[1:])

    return {
        "years": years,
        "wacc": find_wacc(ws),
        "price_base": price_base,
        "price_unit": price_unit,
        "basis_notes": basis,
        "banner1": " | ".join(banner1),
        "banner2": " | ".join(banner2),
        "dcf_years": dcf,
        "metrics": metrics,
    }


# ════════════════════════════════════════════════════════════════════
# Exec Summary — keep every row verbatim (narrative + benchmark + actions)
# ════════════════════════════════════════════════════════════════════
def parse_exec_summary(ws) -> tuple[str, str, list[dict]]:
    headline = clean(ws.cell(row=1, column=1).value)
    subtitle = clean(ws.cell(row=2, column=1).value)
    rows: list[dict] = []
    section = None
    for i in range(3, ws.max_row + 1):
        a = clean(ws.cell(row=i, column=1).value)
        if a is None:
            continue
        # numbered section header like "1.  KEY ASSUMPTIONS"
        if re.match(r"^\d+\.\s", a) and not any(clean(ws.cell(row=i, column=c).value) for c in range(2, 6)):
            section = a
            continue
        rows.append({
            "row_index": i,
            "section": section,
            "metric": a,
            "col_b": clean(ws.cell(row=i, column=2).value),
            "col_c": clean(ws.cell(row=i, column=3).value),
            "col_d": clean(ws.cell(row=i, column=4).value),
            "col_e": clean(ws.cell(row=i, column=5).value),
            "notes": clean(ws.cell(row=i, column=6).value),
        })
    return headline, subtitle, rows


# ════════════════════════════════════════════════════════════════════
# Mine-level metadata from the Exec Summary banners
# ════════════════════════════════════════════════════════════════════
def parse_mine_meta(headline: str, subtitle: str, source_file: str) -> dict:
    parts = split_banner(headline) + split_banner(subtitle)
    blob = " | ".join(parts)

    name = parts[0].title() if parts else "Unknown Mine"
    lic  = (re.search(r"License\s+([0-9A-Za-z]+)", blob) or [None, None])[1]
    prov = (re.search(r"([A-Za-zçè'’]+)\s+Province", blob) or [None, None])[1]

    def grab(pattern, cast=float):
        m = re.search(pattern, blob)
        return cast(m.group(1)) if m else None

    ore        = grab(r"([\d.]+)\s*M(?:t|m³)")
    reserve_u  = (re.search(r"[\d.]+\s*(M[t]|Mm³|m³)", blob) or [None, None])[1]
    lom        = grab(r"(\d+)\s*-?\s*[Yy]ear\s+LOM", int)
    wacc       = grab(r"(\d+(?:\.\d+)?)\s*%\s*WACC")
    tax        = grab(r"(\d+(?:\.\d+)?)\s*%\s*IRPC")
    area_ha    = grab(r"([\d,]+)\s*ha", lambda s: float(s.replace(",", "")))
    tput_m     = re.search(r"([\d.,]+)\s*(Mtpa|m³/yr)", blob)
    throughput = float(tput_m.group(1).replace(",", "")) if tput_m else None
    throughput_u = tput_m.group(2) if tput_m else None

    return {
        "mine_name": name,
        "license_number": lic or source_file,
        "province": prov,
        "ore_reserve": ore,
        "reserve_unit": reserve_u,
        "concession_area_ha": area_ha,
        "throughput_pa": throughput,
        "throughput_unit": throughput_u,
        "life_of_mine_yr": lom,
        "wacc": wacc / 100.0 if wacc else None,
        "tax_rate": tax / 100.0 if tax else None,
        "headline": headline,
        "subtitle": subtitle,
        "source_file": source_file,
    }


def backfill_mine_meta(mine: dict, commodities: dict):
    """Fill ore/throughput/LOM/WACC from commodity sheet banners + dcf length."""
    # WACC from any sheet
    if mine["wacc"] is None:
        for c in commodities.values():
            for s in c["scenarios"]:
                if s["wacc"]:
                    mine["wacc"] = s["wacc"]
                    break
    # LOM = max year index seen across all scenarios
    max_yr = 0
    tput_from_sheet = None
    tput_unit_from_sheet = None
    for c in commodities.values():
        for s in c["scenarios"]:
            if s["years"]:
                max_yr = max(max_yr, max(s["years"]))
            # primary commodity steady-state production hints at throughput
            for raw in (s.get("price_unit"),):
                pass
    if mine["life_of_mine_yr"] is None and max_yr:
        mine["life_of_mine_yr"] = max_yr
    # throughput / ore from any sheet banner (primary commodity preferred)
    if mine["throughput_pa"] is None:
        for c in commodities.values():
            for s in c["scenarios"]:
                blob = (s.get("banner1", "") + " | " + s.get("banner2", ""))
                tm = re.search(r"([\d.,]+)\s*(Mtpa|m³/yr|t/yr)", blob)
                if tm:
                    mine["throughput_pa"] = float(tm.group(1).replace(",", ""))
                    mine["throughput_unit"] = tm.group(2)
                    break
            if mine["throughput_pa"]:
                break


# ════════════════════════════════════════════════════════════════════
# Classify each sheet → (commodity, scenario)
# ════════════════════════════════════════════════════════════════════
def classify_sheet(name: str) -> Optional[tuple[str, str, bool]]:
    """Returns (commodity, scenario, has_scenarios) or None to skip."""
    n = name.strip()
    low = n.lower()
    if low in {"exec summary", "scenario analysis", "ree scenario analysis", "combined dcf"}:
        return None
    m = re.match(r"graphite\s+(bear|base|bull)$", low)
    if m:
        return ("Graphite", m.group(1).capitalize(), True)
    # single-scenario commodity sheets: REE, Monazite, Spodumene, Gold
    return (n, "Single", False)


# ════════════════════════════════════════════════════════════════════
# Build the full payload for one workbook
# ════════════════════════════════════════════════════════════════════
def parse_workbook(path: str) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    source_file = os.path.basename(path)

    headline, subtitle, exec_rows = parse_exec_summary(wb["Exec Summary"])
    mine = parse_mine_meta(headline, subtitle, source_file)

    commodities: dict[str, dict] = {}
    for sheet in wb.sheetnames:
        cls = classify_sheet(sheet)
        if cls is None:
            continue
        commodity, scenario, has_scen = cls
        parsed = parse_dcf_sheet(wb[sheet])
        commodities.setdefault(commodity, {"has_scenarios": has_scen, "scenarios": []})
        commodities[commodity]["scenarios"].append({
            "scenario": scenario,
            "sheet_name": sheet,
            **parsed,
        })

    # primary commodity = the one with the largest positive NPV
    def npv_of(c):
        return max((s["metrics"].get("npv") or -1e18) for s in c["scenarios"])
    primary = max(commodities, key=lambda k: npv_of(commodities[k])) if commodities else None
    for k, c in commodities.items():
        c["is_primary"] = (k == primary)

    backfill_mine_meta(mine, commodities)

    return {"mine": mine, "exec_rows": exec_rows, "commodities": commodities}


# ════════════════════════════════════════════════════════════════════
# Supabase writer
# ════════════════════════════════════════════════════════════════════
def push_to_supabase(payload: dict):
    from supabase import create_client
    # url = os.environ["SUPABASE_URL"]
    # key = os.environ["SUPABASE_SERVICE_KEY"]
    url = os.environ.get("SUPABASE_URL","https://snbnqwrxvptrfjsecljd.supabase.co")
    key = os.environ.get("SUPABASE_SERVICE_KEY",
    "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9."
    "eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InNuYm5xd3J4dnB0cmZqc2VjbGpkIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3OTg5MzQ4MSwiZXhwIjoyMDk1NDY5NDgxfQ."
    "t92K9HW0uQpCWy08c6CPtGKynHN5ET3ymcfcupNJTO0")
    sb = create_client(url, key)

    mine = payload["mine"]

    # upsert mine on license_number
    sb.table("mines").upsert(mine, on_conflict="license_number").execute()
    mine_id = sb.table("mines").select("id").eq(
        "license_number", mine["license_number"]).single().execute().data["id"]

    # wipe child rows for idempotent re-ingest
    sb.table("financial_models").delete().eq("mine_id", mine_id).execute()
    sb.table("mine_commodities").delete().eq("mine_id", mine_id).execute()
    sb.table("exec_summary_rows").delete().eq("mine_id", mine_id).execute()

    # financial model
    fm = sb.table("financial_models").insert({
        "mine_id": mine_id,
        "wacc": mine["wacc"],
        "tax_rate": mine["tax_rate"],
        "source_file": mine["source_file"],
    }).execute().data[0]

    # exec rows
    for r in payload["exec_rows"]:
        sb.table("exec_summary_rows").insert({"mine_id": mine_id, **r}).execute()

    # commodities → scenarios → metrics + dcf_years
    for commodity, cdata in payload["commodities"].items():
        mc = sb.table("mine_commodities").insert({
            "mine_id": mine_id,
            "commodity": commodity,
            "is_primary": cdata["is_primary"],
            "has_scenarios": cdata["has_scenarios"],
        }).execute().data[0]

        for s in cdata["scenarios"]:
            cs = sb.table("commodity_scenarios").insert({
                "commodity_id": mc["id"],
                "scenario": s["scenario"],
                "sheet_name": s["sheet_name"],
                "wacc": s["wacc"],
                "price_base": s["price_base"],
                "price_unit": s["price_unit"],
                "basis_notes": s["basis_notes"],
            }).execute().data[0]
            m = s["metrics"]
            sb.table("scenario_metrics").insert({
                "scenario_id": cs["id"],
                "npv": m.get("npv"), "irr": m.get("irr"), "payback": m.get("payback"),
                "moic": m.get("moic"), "total_capex": m.get("total_capex"),
                "total_lom_fcf": m.get("total_lom_fcf"),
            }).execute()

            rows = [{"scenario_id": cs["id"], **d} for d in s["dcf_years"]]
            if rows:
                sb.table("dcf_years").insert(rows).execute()

    print(f"  ✓ ingested {mine['mine_name']} ({mine['license_number']})")


# ════════════════════════════════════════════════════════════════════
# CLI
# ════════════════════════════════════════════════════════════════════
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("files", nargs="+")
    ap.add_argument("--dry-run", action="store_true", help="parse + summarize, write nothing")
    ap.add_argument("--json", metavar="OUT", help="dump parsed payloads to a JSON file")
    args = ap.parse_args()

    payloads = []
    for f in args.files:
        print(f"Parsing {os.path.basename(f)} …")
        p = parse_workbook(f)
        payloads.append(p)
        m = p["mine"]
        print(f"  mine: {m['mine_name']} | license {m['license_number']} | "
              f"WACC {m['wacc']} | LOM {m['life_of_mine_yr']}")
        for c, cd in p["commodities"].items():
            for s in cd["scenarios"]:
                mt = s["metrics"]
                print(f"    {c:10s} [{s['scenario']:6s}] "
                      f"NPV={mt.get('npv')} IRR={mt.get('irr')} "
                      f"payback={mt.get('payback')} years={len(s['dcf_years'])}")

    if args.json:
        with open(args.json, "w") as fh:
            json.dump(payloads, fh, indent=2, default=str)
        print(f"\nWrote {args.json}")

    if args.dry_run or args.json:
        return

    for p in payloads:
        push_to_supabase(p)


if __name__ == "__main__":
    main()