"""
mine_data_supabase.py
=====================
Two-mode module:

  RUNTIME (imported by api_mines.py)
  ─────────────────────────────────
  get_all_mines()  → list of mine dicts (from v1_mines)
  get_mine(id)     → one mine dict or None
  upsert_mine(id, data) → insert or update in v1_mines

  INGESTION (run as __main__)
  ─────────────────────────────────
  Reads 'Mine_Profile' sheet from the prototype workbook and writes ONE row
  to v1_mines.  Only C-column input cells are read; Calcs sheet is ignored
  entirely — the DCF engine recalculates it at runtime.

  Usage:
    python mine_data_supabase.py "Mining Investment Model Prototype (4.20.2026).xlsx"
    python mine_data_supabase.py --dry-run "..."
"""
from __future__ import annotations
import os, re, sys, json
from typing import Any, Dict, List, Optional

SUPABASE_URL = os.environ.get("SUPABASE_URL",
    "https://snbnqwrxvptrfjsecljd.supabase.co")
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_KEY",
    os.environ.get("SUPABASE_ANON_KEY",
    "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9."
    "eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InNuYm5xd3J4dnB0cmZqc2VjbGpkIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3OTg5MzQ4MSwiZXhwIjoyMDk1NDY5NDgxfQ."
    "t92K9HW0uQpCWy08c6CPtGKynHN5ET3ymcfcupNJTO0"))


def _sb():
    from supabase import create_client
    return create_client(SUPABASE_URL, SUPABASE_KEY)


# ═══════════════════════════════════════════════════════════════════
# RUNTIME  — used by api_mines.py
# ═══════════════════════════════════════════════════════════════════

def get_all_mines() -> List[Dict]:
    """Return all v1_mines rows, sorted by mine_name."""
    res = _sb().table("v1_mines").select("*").order("mine_name").execute()
    return res.data or []


def get_mine(mine_id: str) -> Optional[Dict]:
    """Return one mine dict or None."""
    res = _sb().table("v1_mines").select("*").eq("id", mine_id).limit(1).execute()
    return res.data[0] if res.data else None


def upsert_mine(mine_id: Optional[str], data: Dict) -> Dict:
    """
    Insert (mine_id=None) or update an existing row.
    Returns the saved row.
    """
    client = _sb()
    if mine_id:
        client.table("v1_mines").update(data).eq("id", mine_id).execute()
        return get_mine(mine_id)
    else:
        res = client.table("v1_mines").insert(data).execute()
        return res.data[0]


def _summary(mine: Dict) -> Dict:
    """Lightweight summary dict for list endpoints."""
    return {
        "id":              mine.get("id"),
        "mine_name":       mine.get("mine_name"),
        "license_number":  mine.get("license_number"),
        "commodity":       mine.get("commodity"),
        "province":        mine.get("province"),
        "wacc":            mine.get("wacc"),
        "tax_rate":        mine.get("tax_rate"),
        "life_of_mine_yr": mine.get("life_of_mine_yr"),
        "status":          mine.get("status"),
    }


# ═══════════════════════════════════════════════════════════════════
# INGESTION  — parse Mine_Profile sheet and write to v1_mines
# ═══════════════════════════════════════════════════════════════════

# Row → field mapping for Mine_Profile sheet.
# Verified against: Mining Investment Model Prototype (4.20.2026).xlsx
# Only INPUT rows — computed rows (21-22, 29-30, 49-58) are skipped.
# Format: (row_number, col, field_name, transform_fn)
#   col=3 → column C (value), col=2 → column B (unit label)
_PROFILE_MAP = [
    # Identity
    (4,  3, "mine_name",                 str),
    (6,  3, "country",                   str),
    (7,  3, "license_number",            str),
    (8,  3, "mine_type",                 str),
    # Reserve & throughput
    (13, 3, "ore_reserve",               float),
    (13, 2, "reserve_unit",              str),   # col B = "m3"
    (14, 3, "throughput_pa",             float),
    (15, 3, "life_of_mine_yr",           int),
    # Commodity & production (annual_production row 21 is computed — skip)
    (18, 3, "commodity",                 str),
    (19, 3, "price_base",                float),
    (20, 3, "grade",                     float),
    (20, 2, "grade_unit",                str),   # col B = "g/m3"
    # Capital & operating costs
    (25, 3, "initial_dev_capex",         float),
    (28, 3, "total_opex_steady_state",   float),
    (31, 3, "opex_escalation_rate",      float),
    (32, 3, "avg_depreciation_years",    int),
    # Ramp-up
    (35, 3, "ramp_up_y1",                float),
    (36, 3, "ramp_up_y2",                float),
    (37, 3, "ramp_up_y3",                float),
    # Finance
    (40, 3, "tax_rate",                  float),
    (41, 3, "royalty_rate",              float),
    (42, 3, "debt_funding",              float),
    (43, 3, "debt_term",                 int),
    (44, 3, "interest_rate",             float),
    (45, 3, "wacc",                      float),
    # End of life
    (46, 3, "closure_rehab_cost",        float),
]


def _safe(fn, val):
    if val is None or val == "":
        return None
    try:
        return fn(val)
    except (ValueError, TypeError):
        return None


def parse_profile_sheet(wb_path: str) -> Dict:
    """
    Parse the 'Mine_Profile' sheet from the prototype workbook.
    Returns a dict ready to insert into v1_mines.
    """
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl not installed — run: pip install openpyxl")

    wb = openpyxl.load_workbook(wb_path, data_only=True)

    # Find sheet — try several name variants
    sheet = None
    for name in ["Mine_Profile", "Mine Profile", "MineProfile", "Profile"]:
        if name in wb.sheetnames:
            sheet = wb[name]
            break
    if sheet is None:
        raise ValueError(f"No Mine_Profile sheet found. Available: {wb.sheetnames}")

    row_vals = {}
    for row_num, col, field, fn in _PROFILE_MAP:
        cell_val = sheet.cell(row=row_num, column=col).value
        row_vals[field] = _safe(fn, cell_val)

    # Remove None values so Supabase defaults apply
    data = {k: v for k, v in row_vals.items() if v is not None}
    data.setdefault("country", "Mozambique")
    data.setdefault("is_user_created", False)
    data["source_file"] = os.path.basename(wb_path)
    return data


def ingest_workbook(wb_path: str, dry_run: bool = False) -> Dict:
    """
    Parse Mine_Profile sheet and insert into v1_mines.
    Returns the inserted row (or parsed dict on dry_run).
    """
    data = parse_profile_sheet(wb_path)
    if dry_run:
        print(json.dumps(data, indent=2, default=str))
        return data

    mine_name = data.get("mine_name", "Unknown")
    license   = data.get("license_number")

    # Upsert by license_number if it already exists
    existing = None
    if license:
        res = _sb().table("v1_mines").select("id").eq("license_number", license).limit(1).execute()
        existing = res.data[0]["id"] if res.data else None

    result = upsert_mine(existing, data)
    print(f"{'Updated' if existing else 'Inserted'}: {mine_name} (id={result.get('id')})")
    return result


# ═══════════════════════════════════════════════════════════════════
# CLI entry point
# ═══════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Ingest Mine_Profile into v1_mines")
    parser.add_argument("workbook", nargs="+", help="Excel workbook path(s)")
    parser.add_argument("--dry-run", action="store_true", help="Parse only, no DB writes")
    args = parser.parse_args()

    for wb in args.workbook:
        print(f"\n── {wb} ──")
        try:
            ingest_workbook(wb, dry_run=args.dry_run)
        except Exception as e:
            print(f"  ERROR: {e}", file=sys.stderr)
